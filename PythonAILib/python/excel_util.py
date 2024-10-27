import sys
import datetime
sys.path.append("python")

import openpyxl


def export_to_excel(filePath, data):
    # Workbookオブジェクトを生成
    wb = openpyxl.Workbook()
    # アクティブなシートを取得
    ws = wb.active
    # シート名を設定
    ws.title = "Sheet1"
    # データを書き込む
    for row in data:
        ws.append(row)
    # ファイルを保存
    wb.save(filePath)
    
def import_from_excel(filePath):
    # Workbookオブジェクトを生成
    wb = openpyxl.load_workbook(filePath)
    # アクティブなシートを取得
    ws = wb.active
    # データを取得
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)
    
    return data

# application/vnd.openxmlformats-officedocument.spreadsheetml.sheetのファイルを読み込んで文字列として返す関数
def extract_text_from_sheet(filename:str, sheet_name:str=""):
    import openpyxl
    from io import StringIO
    # 出力用のストリームを作成
    output = StringIO()
    wb = openpyxl.load_workbook(filename)
    for sheet in wb:
        # シート名が指定されている場合はそのシートのみ処理
        if sheet_name and sheet.title != sheet_name:
            continue
        for row in sheet.iter_rows(values_only=True):
            # 1行分のデータを格納するリスト
            cells = []
            for cell in row:
                # cell.valueがNoneの場合はcontinue
                if cell is None:
                    continue
                # cell.valueがdatetime.datetimeの場合はisoformat()で文字列に変換
                if isinstance(cell, datetime.datetime):
                    cells.append(cell.isoformat())
                else:
                    cells.append(str(cell))
                
            output.write("\t".join(cells))
            output.write("\n")
    
    return output.getvalue()

# excelのシート名一覧を取得する関数
def get_sheet_names(filename):
    import openpyxl
    wb = openpyxl.load_workbook(filename)
    return wb.sheetnames

