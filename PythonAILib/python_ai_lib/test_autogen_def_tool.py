import sys
from openpyxl import load_workbook

def create_function(content: str):
    # contentから関数オブジェクトを作成する。
    exec_locals = {}
    exec(content, {}, exec_locals)
    function_name = list(exec_locals.keys())[0]  # 1つ目の関数名を取得
    return exec_locals[function_name]

def main(excel_file):
    # Excelファイルを開く
    wb = load_workbook(excel_file)
    sheet = wb['tools']

    # ヘッダー行から列名を取得
    headers = [cell.value for cell in sheet[1]]
    name_index = headers.index('name')
    description_index = headers.index('description')
    content_index = headers.index('content')

    function_dict = {}

    # 2行目から最終行まで処理を行う
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = {headers[i]: row[i] for i in range(len(headers))}
        
        name = row_data['name']  # name列
        description = row_data['description']  # description列
        content = row_data['content']  # content列

        # 関数を作成
        function_obj = create_function(content)

        # dictに格納
        function_dict[name] = (function_obj, description)

    # 結果を表示（必要に応じて）
    for name, (func, desc) in function_dict.items():
        print(f'Name: {name}, Description: {desc}, Function: {func}')

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python script.py <excel_file>")
        sys.exit(1)

    excel_file = sys.argv[1]
    main(excel_file)