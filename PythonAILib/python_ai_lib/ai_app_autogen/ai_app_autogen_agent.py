
from autogen import ConversableAgent, UserProxyAgent
import autogen 
from collections.abc import Generator
from typing import Any

from ai_app_autogen.ai_app_autogen_tools import AutoGenTools, AutoGenToolGenerator
from ai_app_autogen.ai_app_autogen_client import AutoGenProps
from ai_app_vector_db.ai_app_vector_db_util import VectorDBProps

class AutoGenAgents:

    def __init__(self, autogen_props: AutoGenProps, autogen_tools: AutoGenTools, agents_dict: list[dict], auto_execute_code: bool = False):
        
        self.autogen_props = autogen_props
        self.work_dir_path = autogen_props.work_dir_path
        # コード実行者が自動的にコードを実行するようにするかどうか
        # self.auto_execute_code = auto_execute_code
        self.auto_execute_code = True
        
        self.autogen_tools = autogen_tools

        self.agents : dict[str, tuple[ConversableAgent, str]] = {}
        if self.autogen_props.use_system_agent:
            print("Using system agents")
            self.agents.update(AutoGenAgentGenerator.create_default_agents(self.autogen_props, self.autogen_tools))
        
        for agent_dict in agents_dict:
            self.agents.update(AutoGenAgentGenerator.create_agents_dict(self.autogen_tools, agent_dict))

        self.temp_dir = None


    def add_agents(self, agents: dict[str, tuple[ConversableAgent, str]]):
        self.agents.update(agents)

    # エージェントの終了処理
    def finish(self):
        if self.temp_dir:
            self.temp_dir.cleanup()

from openpyxl import load_workbook
from autogen import ConversableAgent, UserProxyAgent
from ai_app_autogen.ai_app_autogen_client import AutoGenProps

class AutoGenAgentGenerator:

    @classmethod
    def create_agent(cls, autogen_tools: AutoGenTools, params: dict) -> ConversableAgent:
        autogen_props: AutoGenProps = autogen_tools.autogen_props

        # human_input_modeが指定されていない場合はデフォルト値を設定
        human_input_mode = params.get('human_input_mode', None)
        params['human_input_mode'] = human_input_mode if human_input_mode else "NEVER"
        # is_termination_msgが指定されていない場合はNone, それ以外の場合はtermination_msgを設定
        termination_msg = params.get('is_termination_msg', None)
        if termination_msg:
            is_termination_msg = lambda msg: termination_msg in msg["content"].lower()
            params['is_termination_msg'] = is_termination_msg

        # code_executionを大文字にしたものがTRUEの場合はauto_gen_pros.create_code_executor()を呼び出し
        # それ以外の場合はFalseを設定
        code_execution = params.get('code_execution_config', None)
        params['code_execution_config'] = autogen_props.create_code_executor() if code_execution.upper() == "TRUE" else False
        # llm_configを設定 llmをお大文字にしたものがTRUEの場合はauto_gen_pros.create_llm_config()を呼び出し
        # それ以外の場合はNoneを設定
        llm = params.get('llm_config', None)
        params['llm_config'] = autogen_props.create_llm_config() if llm.upper() == "TRUE" else None

        type_value = params.get('type', None)
        agent = None
        if type_value == 'userproxy':
            # userproxyの場合
            agent: UserProxyAgent = UserProxyAgent( **params)
        elif type_value == 'assistant':
            # assistantの場合
            agent: ConversableAgent = ConversableAgent( **params)
        else:
            raise ValueError(f"Unknown agent type: {type_value}")

        # agentにツールを設定
        # tools :str からtools_listに変換
        tool_name_text = params.get('tools', None)
        if tool_name_text:
            tool_names = tool_name_text.split(',')
            for tool_name in tool_names:
                # self.autogen_toolsからtool_nameを指定して関数と説明を取得
                function_obj, description = autogen_tools.get(tool_name, (None, None))
                if function_obj is None:
                    raise ValueError(f"Unknown tool: {tool_name}")
                # agentにツールを設定
                agent.register_for_llm(name=tool_name, description=description)(function_obj)
        
        return agent

    @classmethod
    def create_agents_dict(cls, autogen_tools: AutoGenTools, data_list: list[dict]) -> dict:
        agent_dict = {}
        for data in data_list:
            name = data['name']
            description = data['description']
            # エージェントを作成
            agent_obj = cls.create_agent(autogen_tools, data)
            # dictに格納
            agent_dict[name] = (agent_obj, description)

        # 結果を表示（必要に応じて）
        for name, (agent, desc) in agent_dict.items():
            print(f'Name: {name}, Description: {desc}, Agent: {agent}')
        
        return agent_dict

    @classmethod
    def create_agents_dict_from_sheet(cls, autogen_tools: AutoGenTools, excel_file):
        # Excelファイルを開く
        wb = load_workbook(excel_file)
        sheet = wb['agents']

        # ヘッダー行から列名を取得
        headers = [cell.value for cell in sheet[1]]
        name_index = headers.index('name')
        description_index = headers.index('description')
        system_message_index = headers.index('system_message')
        type_index = headers.index('type')
        tools_index = headers.index('tools')
        human_input_mode_index = headers.index('human_input_mode')
        is_termination_msg_index = headers.index('is_termination_msg')
        code_execution_config_index = headers.index('code_execution_config')
        llm_config_index = headers.index('llm_config')

        agent_data_list = []
        # 2行目から最終行まで処理を行う
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # 各列の値を取得
            name = row[name_index]
            description = row[description_index]
            system_message = row[system_message_index]
            type_value = row[type_index]
            tools = row[tools_index]
            human_input_mode = row[human_input_mode_index]
            is_termination_msg = row[is_termination_msg_index]
            code_execution_config = row[code_execution_config_index]
            llm_config = row[llm_config_index]

            agent_data_list.append({
                'name': name, 'description': description, 'system_message': system_message, 
                'type': type_value, 'tools': tools, 'human_input_mode': human_input_mode,
                'is_termination_msg': is_termination_msg, 'code_execution_config': code_execution_config,
                'llm_config': llm_config
            })
            return cls.create_agents_dict(autogen_tools, agent_data_list)
    
    @classmethod
    def create_default_agents(cls, autogen_props: AutoGenProps, tools: AutoGenTools) -> dict[str, tuple[ConversableAgent, str]]:

        agents : dict[str, tuple[ConversableAgent, str]] = {}
        agents["user_proxy"] = cls.__create_user_proxy(autogen_props, tools)
        agents["code_executor"] = cls.__create_code_executor(autogen_props, tools, True)
        agents["code_writer"] = cls.__create_code_writer(autogen_props, tools)
        agents["file_writer"] = cls.__create_file_writer(autogen_props, tools)
        agents["vector_searcher"] = cls.__create_vector_searcher(autogen_props, tools)
        agents["file_extractor"] = cls.__create_file_extractor(autogen_props, tools)
        agents["web_searcher"] = cls.__create_web_searcher(autogen_props, tools)
        agents["wikipeda_searcher"] = cls.__create_wikipedia_searcher(autogen_props, tools)
        agents["azure_document_searcher"] = cls.__create_azure_document_searcher(autogen_props, tools)
        agents["file_checker"] = cls.__create_file_checker(autogen_props, tools)
        agents["current_time"] = cls.__create_current_time(autogen_props, tools)

        return agents

    @classmethod
    def __create_user_proxy(cls, autogen_pros: AutoGenProps, autogen_tools: AutoGenTools):
        # グループチャットの議題提供者
        description = "ユーザーの依頼を達成するためのタスク一覧を考えて各エージェントにタスクを割り当てます。"
        user_proxy = autogen.UserProxyAgent(
            system_message="""
                ユーザーの依頼を達成するため各エージェントと協力してタスクを実行します。
                - まず、ユーザーの依頼を達成するためのプランを考えて、タスク一覧を作成します。
                - 各エージェントにタスクを割り当て、タスクを実行します。
                - 各エージェントによるタスク実行によりプランが達成されたら、[会議を終了]と返信します。
                - 追加質問がない場合は、[会議を終了]と返信します。
                """,
            name="user_proxy",
            human_input_mode="NEVER",
            code_execution_config=False,
            is_termination_msg=lambda msg: "会議を終了" in msg["content"].lower(),
            description=description,
            llm_config=autogen_pros.create_llm_config()
        )

        # user_proxyのregister_from_execution()を設定
        for func, description  in autogen_tools.tools.values():
            print(f"register_for_execution: {func.__name__}")
            user_proxy.register_for_execution()(func)

        return user_proxy, description

    @classmethod
    def __create_code_writer(cls, autogen_pros: AutoGenProps, autogen_tools: AutoGenTools):
        # コードの作成者と実行者は分離する。以下はコード推論 Agent。LLM を持つ。
        description = "ユーザーの指示に従い、Pythonスクリプトを作成します。"
        code_writer_agent = ConversableAgent(
            "code-writer",
            system_message=f"""
                あなたはスクリプト開発者です。
                あなたがコードを書くと自動的に外部のアプリケーション上で実行されます。
                ユーザーの指示に従ってあなたはコードを書きます。
                コードの実行結果は、あなたがコードを投稿した後に自動的に表示されます。
                ただし、次の条件を厳密に遵守する必要があります。
                ルール:
                * コードブロック内でのみコードを提案します。
                * スクリプトの実行結果がエラーである場合、エラー文を元に対策を考え、修正したコードを再度作成します。
                * スクリプトを実行した結果、情報を十分に得られなかった場合、現状得られた情報から修正したコードを再度作成します。
                * あなたはユーザーの指示を最終目標とし、これを満たす迄何度もコードを作成・修正します。
                """
        ,
            llm_config=autogen_pros.create_llm_config(),
            code_execution_config=False,
            description=description,
            human_input_mode="NEVER",
        )

        return code_writer_agent, description

    @classmethod
    def __create_file_writer(cls, autogen_pros: AutoGenProps, autogen_tools: AutoGenTools):
        # コードの作成者と実行者は分離する。以下はコード推論 Agent。LLM を持つ。
        description = "ユーザーの指示に従い、Pythonでデータをファイルに保存します。"
        file_writer = ConversableAgent(
            "file-writer",
            system_message=f"""
                ユーザーの指示に従い、Pythonでデータをファイルに保存します。
                デフォルトの保存場所は{autogen_pros.work_dir_path}です。
                ユーザーからファイルの保存場所の指定があれば、指定された場所にファイルを保存します。
                """
        ,
            llm_config=autogen_pros.create_llm_config(),
            code_execution_config=False,
            description=description,
            human_input_mode="NEVER",
        )
        save_tools, description = autogen_tools.tools["save_text_file"]
        # register_for_llm
        file_writer.register_for_llm(description=description)(save_tools)

        return file_writer, description

    @classmethod
    def __create_code_executor(cls, autogen_pros: AutoGenProps, autogen_tools: AutoGenTools, auto_execute_code: bool = False):
        # コードの作成者と実行者は分離する。以下はコード実行者 Agent。LLM は持たない。
        description = "コード作成者から提供されたコードを実行します。"
        code_execution_agent = ConversableAgent(
            "code-execution",
            system_message=f"""
                あなたはコード実行者です。
                コード作成者から提供されたコードを実行します。
                コードの実行結果を表示します。
                """,
            llm_config=False,
            code_execution_config={"executor": autogen_pros.create_code_executor()},   
            description=description,
            human_input_mode="NEVER",
        )

        # コード実行者が自動的にコードを実行するようにするかどうか
        # self.auto_execute_code == True の場合はhuman_input_mode="NEVER"とする
        if auto_execute_code:
            code_execution_agent.human_input_mode = "NEVER"
        else:
            code_execution_agent.human_input_mode = "ALWAYS" 

        return code_execution_agent, description

    # ベクトル検索者を有効にする
    @classmethod
    def __create_vector_searcher(cls, autogen_pros: AutoGenProps, autogen_tools: AutoGenTools):
        # ベクトル検索者
        description = "ユーザーの指示に従い、ベクトルデータベースから情報を検索します。"
        vector_searcher = ConversableAgent(
            "vector-searcher",
            system_message="""
                あなたはベクトル検索者です。ユーザーの指示に従い、ベクトルデータベースから情報を検索します。
                提供された関数を用いて、検索した結果を表示します。
                """,
            llm_config=autogen_pros.create_llm_config(),
            code_execution_config=False,
            description=description,
            human_input_mode="NEVER",
        )

        # 利用可能な関数の情報をエージェントに登録する
        vector_search, description = autogen_tools.tools["vector_search"]
        vector_searcher.register_for_llm(description=description)(vector_search)
        
        return vector_searcher, description

    # ファイル抽出者を有効にする
    @classmethod
    def __create_file_extractor(cls, autogen_pros: AutoGenProps, autogen_tools: AutoGenTools):
        # ファイル抽出者
        description = "ユーザーの指示に従い、ファイルから情報を抽出します。"
        file_extractor = ConversableAgent(
            "file-extractor",
            system_message="""
                あなたはファイル抽出者です。ユーザーの指示に従い、ファイルから情報を抽出します。
                提供された関数を用いて、抽出した結果を表示します。
                """,
            llm_config=autogen_pros.create_llm_config(),
            code_execution_config=False,
            description=description,
            human_input_mode="NEVER",
        )

        # 利用可能な関数の情報をエージェントに登録する
        extract_text_from_file, description = autogen_tools.tools["extract_text_from_file"]
        file_extractor.register_for_llm (description=description)(extract_text_from_file)

        list_files_in_directory,  description = autogen_tools.tools["list_files_in_directory"]
        file_extractor.register_for_llm (description=description)(list_files_in_directory)

        return file_extractor, description

    @classmethod
    def __create_web_searcher(cls, autogen_pros: AutoGenProps, autogen_tools: AutoGenTools):
        # Web検索者
        description = "Webのドキュメントを検索します。"
        web_searcher = ConversableAgent(
            "azure-document-searcher",
            system_message="""
                あなたはWeb検索者です。ユーザーの指示に従いWeb上のドキュメントを検索します。
                - 提供されたsearch_duckduckgo関数を用いて、情報を検索します。
                - リンク先に必要なドキュメントがない場合はさらにリンクされた情報を検索します。
                - 必要なドキュメントがあった場合はドキュメントをextract_webpageで取得して、テキストをユーザーに提供します。
                """,
            llm_config=autogen_pros.create_llm_config(),
            code_execution_config=False,
            description=description,
            human_input_mode="NEVER",
        )

        # 指定したキーワードでAzureのドキュメントを検索します。
        search_duckduckgo, description = autogen_tools.tools["search_duckduckgo"]
        web_searcher.register_for_llm(description=description)(search_duckduckgo)

        # 指定されたURLのテキストとリンクを取得します。
        extract_webpage, description = autogen_tools.tools["extract_webpage"]
        web_searcher.register_for_llm(description=description)(extract_webpage)

        return web_searcher, description

    @classmethod
    def __create_wikipedia_searcher(cls, autogen_pros: AutoGenProps, autogen_tools: AutoGenTools):
        
        # Web検索者
        description = "指定されたURLから情報を取得します。"
        wipkipedia_searcher = ConversableAgent(
            "wipkipedia-searcher",
            system_message="""
                ユーザーの指示に従い、指定されたURLから情報を取得します。
                - 提供された関数を用いて、指定されたURLのテキストとリンクを取得します。
                - ユーザーからURLが提供されなかった場合は、WikiPedia(日本語版)から検索対象文字列に関連するページを検索します。
                - リンク先に必要なドキュメントがない場合はさらにリンクされた情報を検索します。
                - 必要なドキュメントがあった場合はドキュメントのテキストをユーザーに提供します
                """,
            llm_config=autogen_pros.create_llm_config(),
            code_execution_config=False,   
            description=description,
            human_input_mode="NEVER",
        )

        # 利用可能な関数の情報をエージェントに登録する

        # Wikipedia(日本語版)から検索対象文字列に関連するページを検索します。
        search_wikipedia_ja, description = autogen_tools.tools["search_wikipedia_ja"]
        wipkipedia_searcher.register_for_llm(description=description)(search_wikipedia_ja)

        # 指定されたURLのテキストとリンクを取得します。
        extract_webpage, description = autogen_tools.tools["extract_webpage"]
        wipkipedia_searcher.register_for_llm(description=description)(extract_webpage)

        return wipkipedia_searcher, description

    @classmethod
    def __create_azure_document_searcher(cls, autogen_pros: AutoGenProps, autogen_tools: AutoGenTools):
        
        # Web検索者
        description = "Azure関連のドキュメントを検索します。"
        azure_document_searcher = ConversableAgent(
            "azure-document-searcher",
            system_message="""
                あなたはAzure関連のドキュメント検索者です。ユーザーの指示に従いWeb上のAzure関連のドキュメントを検索します。
                - 提供されたsearch_duckduckgo関数を用いて、site:https://learn.microsoft.com/ja-jp/azure内の情報を検索します。
                - リンク先に必要なドキュメントがない場合はさらにリンクされた情報を検索します。
                - 必要なドキュメントがあった場合はドキュメントをextract_webpageで取得して、テキストをユーザーに提供します。
                """,
            llm_config=autogen_pros.create_llm_config(),
            code_execution_config=False,
            description=description,
            human_input_mode="NEVER",
        )

        # 指定したキーワードでAzureのドキュメントを検索します。
        search_duckduckgo, description = autogen_tools.tools["search_duckduckgo"]
        azure_document_searcher.register_for_llm(description=description)(search_duckduckgo)

        # 指定されたURLのテキストとリンクを取得します。
        extract_webpage, description = autogen_tools.tools["extract_webpage"]
        azure_document_searcher.register_for_llm(description=description)(extract_webpage)

        return azure_document_searcher, description

    @classmethod
    def __create_file_checker(cls, autogen_pros: AutoGenProps, autogen_tools: AutoGenTools):
        
        # ファイルチェッカー
        description = "指定されたファイルが存在するか確認します。"
        file_checker = ConversableAgent(
            "file_checker",
            system_message="""
                ファイルチェッカーです。指定されたファイルが存在するか確認します。
                """,
            llm_config=autogen_pros.create_llm_config(),
            code_execution_config=False,
            description=description,
            human_input_mode="NEVER",
        )

        # 指定されたURLのテキストとリンクを取得します。
        func, description = autogen_tools.tools["check_file"]
        file_checker.register_for_llm(description=description)(func)

        return file_checker, description

    # 現在の時刻を取得するエージェントを作成
    @classmethod
    def __create_current_time(cls, autogen_pros: AutoGenProps, autogen_tools: AutoGenTools):
        # 現在の時刻を取得するエージェント
        description = "現在の時刻を取得します。"
        current_time = ConversableAgent(
            "current_time",
            system_message="""
                現在の時刻を取得します。
                """,
            llm_config=autogen_pros.create_llm_config(),
            code_execution_config=False,
            description=description,
            human_input_mode="NEVER",
        )

        # 現在の時刻を取得する関数を登録
        func, description = autogen_tools.tools["get_current_time"]
        current_time.register_for_llm(description=description)(func)

        return current_time, description
